from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from collections import Counter
from openpyxl.styles import PatternFill
from openpyxl.formula.translate import Translator
from datetime import datetime
from selenium.webdriver.firefox.options import Options
import logging

rodar_debug = False


def get_data():
    try:
        options = Options()
        options.headless = True
        driver = webdriver.Firefox(options=options, executable_path='C://geckodriver')
        driver.get("https://membros.sunoresearch.com.br/carteira-recomendada-dividendos/")
        # time.sleep(5)
        elem = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "user_email"))
        )
        elem.send_keys("marcosmartin1998@hotmail.com")
        # time.sleep(5)
        elem = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "user_password"))
        )
        elem.send_keys("Marcos24")
        # time.sleep(5)
        elem = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "login_button"))
        )
        elem.click()
        time.sleep(10)
        logging.info(str(datetime.now()) + ": Logou!")
        print(str(datetime.now()) + ": Logou!")
        driver.get("https://membros.sunoresearch.com.br/carteira-recomendada-dividendos/")
        elem_list = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#table_1 > tbody:nth-child(2)"))
        )
        elem_list = elem_list.find_elements_by_tag_name("tr")
        final_list_acoes = []
        final_list_fii = []
        for element in elem_list:
            #verificando se é a linha de renda fixa
            if 'span' not in element.find_element_by_class_name("column-ticker").get_attribute('innerHTML'):
                continue
            else:
                final_list_acoes.append((str(element.find_element_by_class_name("column-ticker").find_element_by_class_name("ticker").find_elements_by_tag_name("span")[0].get_attribute('innerHTML')).strip(), float(element.find_element_by_class_name("column-preo-teto").get_attribute('innerHTML').strip().replace(",","."))))
        
        logging.info(str(datetime.now()) + ": Carteira de dividendos verificada!")
        print(str(datetime.now()) + ": Carteira de dividendos verificada!")
        driver.get("https://membros.sunoresearch.com.br/carteira-recomendada-valor/")
        elem_list = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#table_1 > tbody:nth-child(2)"))
        )
        elem_list = elem_list.find_elements_by_tag_name("tr")
        for element in elem_list: 
            if 'span' not in element.find_element_by_class_name("column-ticker").get_attribute('innerHTML'):
                continue
            else:
                final_list_acoes.append((str(element.find_element_by_class_name("column-ticker").find_element_by_class_name("ticker").find_elements_by_tag_name("span")[0].get_attribute('innerHTML')).strip(), float(element.find_element_by_class_name("column-preo-teto").get_attribute('innerHTML').strip().replace(",","."))))

        logging.info(str(datetime.now()) + ": Carteira de valor verificada!")
        print(str(datetime.now()) + ": Carteira de valor verificada!")

        
        driver.get("https://membros.sunoresearch.com.br/carteira-recomendada-fundos-imobiliarios/")
        elem_list = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#table_1 > tbody:nth-child(2)"))
        )
        elem_list = elem_list.find_elements_by_tag_name("tr")
        for element in elem_list:
                final_list_fii.append((str(element.find_element_by_class_name("column-ativo").get_attribute('innerHTML')).strip(), float(element.find_element_by_class_name("column-preo-teto").get_attribute('innerHTML').strip().replace(",","."))))

        logging.info(str(datetime.now()) + ": Carteira de fundos imobiliários verificada!")
        print(str(datetime.now()) + ": Carteira de fundos imobiliários verificada!")
        return final_list_acoes, final_list_fii
    except Exception as e:
        logging.info(str(datetime.now()) + ": Erro no getData: " + str(e))
        raise e

def update_worksheet(lista_suno_acoes, lista_suno_fii):

    try:
        file_name = "C:/Users/marco/Documents/Google Drive/Marcuzão Axelrod Teste.xlsx" if rodar_debug else "C:/Users/marco/Documents/Google Drive/Marcuzão Axelrod.xlsx"
        wb = openpyxl.load_workbook(filename = file_name)
        ws = wb.get_sheet_by_name("Ações")

        list_ativos_planilha = []
        row_fim = 0
        for row in range(2,ws.max_row):
            cell_name_ativo = "{}{}".format("A", row)
            cell_name_atual = "{}{}".format("B", row)
            cell_name_preco = "{}{}".format("C", row)
            cell_name_desc = "{}{}".format("M", row)
            ativo = str(ws[cell_name_ativo].value).split()
            if ativo[0] == "FIM":
                row_fim = row
                break
            preco = float(ws[cell_name_preco].value)
            list_ativos_planilha.append(ativo[0])
            for item in lista_suno_acoes:
                if ativo[0] == item[0] and preco != item[1]:
                    ws[cell_name_preco] = item[1]
                    ws[cell_name_desc] = "Alterou preço teto em " + datetime.today().strftime('%d/%m/%Y')
                    logging.info(str(datetime.now()) + ": Ativo " + ativo[0] + " alterou preço teto")
                    print(str(datetime.now()) + ": Ativo " + ativo[0] + " alterou preço teto")

        list_ativos_suno = [ x[0] for x in lista_suno_acoes ]
        list_precos_suno = [ x[1] for x in lista_suno_acoes ]
        counter_ativos_planilha = Counter(list_ativos_planilha)
        counter_ativos_suno = Counter(list_ativos_suno)
        intersec = counter_ativos_planilha & counter_ativos_suno
        list_vendidos = list(counter_ativos_planilha - intersec)
        list_comprados = list(counter_ativos_suno - intersec)

        for item_comprado in list_comprados:
            cell_name_ativo = "{}{}".format("A", row_fim)
            cell_name_atual = "{}{}".format("B", row_fim)
            cell_name_preco = "{}{}".format("C", row_fim)
            cell_name_desc = "{}{}".format("M", row_fim)
            ws[cell_name_ativo] = item_comprado
            ws[cell_name_atual] = "=GOOGLEFINANCE({})".format(cell_name_ativo)
            ws[cell_name_preco] = list_precos_suno[list_ativos_suno.index(item_comprado)]
            ws[cell_name_desc] = "Comprado pela Suno em "  + datetime.today().strftime('%d/%m/%Y')
            logging.info(str(datetime.now()) + ": Ativo " + item_comprado + " foi comprado")
            print(str(datetime.now()) + ": Ativo " + item_comprado + " foi comprado")
            for coluna_atual in {"D", "F", "G", "H", "I", "L", "J"}:
                celula_cima = "{}{}".format(coluna_atual, row_fim-1)
                celula_atual = "{}{}".format(coluna_atual, row_fim)
                ws[celula_atual] = Translator(ws[celula_cima].value, celula_cima).translate_formula(celula_atual)
            coluna_atual = "E"
            celula_atual = "{}{}".format(coluna_atual, row_fim)
            ws[celula_atual].value = 0
            coluna_atual = "K"
            celula_atual = "{}{}".format(coluna_atual, row_fim)
            ws[celula_atual].value = "=MAX(INDEX(GOOGLEFINANCE({}; \"high\"; $N$2; $O$2); ; 2))".format(cell_name_ativo)
            row_fim += 1
        
        for item_vendido in list_vendidos:
            cell_name_desc = "{}{}".format("M", list_ativos_planilha.index(item_vendido) + 2)
            if ws[cell_name_desc].value is None or ws[cell_name_desc].value.find("Vendido pela Suno") == -1:
                ws[cell_name_desc] = "Vendido pela Suno em " + datetime.today().strftime('%d/%m/%Y')
                logging.info(str(datetime.now()) + ": Ativo " + item_vendido + " foi vendido")
                print(str(datetime.now()) + ": Ativo " + item_vendido + " foi vendido")
            ws[cell_name_desc].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        ws["{}{}".format("A", row_fim)] = "FIM"

        
        ws = wb.get_sheet_by_name("FII")
        list_ativos_planilha = []
        row_fim = 0
        for row in range(2,ws.max_row):
            cell_name_ativo = "{}{}".format("A", row)
            cell_name_atual = "{}{}".format("B", row)
            cell_name_preco = "{}{}".format("C", row)
            cell_name_desc = "{}{}".format("J", row)
            ativo = str(ws[cell_name_ativo].value).split()
            if ativo[0] == "FIM":
                row_fim = row
                break
            preco = float(ws[cell_name_preco].value)
            list_ativos_planilha.append(ativo[0])
            for item in lista_suno_fii:
                if ativo[0] == item[0] and preco != item[1]:
                    ws[cell_name_preco] = item[1]
                    ws[cell_name_desc] = "Alterou preço teto em " + datetime.today().strftime('%d/%m/%Y')
                    logging.info(str(datetime.now()) + ": FII " + ativo[0] + " alterou preço teto")
                    print(str(datetime.now()) + ": FII " + ativo[0] + " alterou preço teto")

        list_ativos_suno = [ x[0] for x in lista_suno_fii ]
        list_precos_suno = [ x[1] for x in lista_suno_fii ]
        counter_ativos_planilha = Counter(list_ativos_planilha)
        counter_ativos_suno = Counter(list_ativos_suno)
        intersec = counter_ativos_planilha & counter_ativos_suno
        list_vendidos = list(counter_ativos_planilha - intersec)
        list_comprados = list(counter_ativos_suno - intersec)

        print("Lista comprados")
        for item_comprado in list_comprados:
            cell_name_ativo = "{}{}".format("A", row_fim)
            cell_name_atual = "{}{}".format("B", row_fim)
            cell_name_preco = "{}{}".format("C", row_fim)
            cell_name_desc = "{}{}".format("J", row_fim)
            ws[cell_name_ativo] = item_comprado
            ws[cell_name_atual] = "=GOOGLEFINANCE({})".format(cell_name_ativo)
            ws[cell_name_preco] = list_precos_suno[list_ativos_suno.index(item_comprado)]
            ws[cell_name_desc] = "Comprado pela Suno em "  + datetime.today().strftime('%d/%m/%Y')
            logging.info(str(datetime.now()) + ": FII " + item_comprado + " foi comprado")
            print(str(datetime.now()) + ": FII " + item_comprado + " foi comprado")
            for coluna_atual in {"D", "F", "G", "H"}:
                celula_cima = "{}{}".format(coluna_atual, row_fim-1)
                celula_atual = "{}{}".format(coluna_atual, row_fim)
                ws[celula_atual] = Translator(ws[celula_cima].value, celula_cima).translate_formula(celula_atual)
            coluna_atual = "E"
            celula_atual = "{}{}".format(coluna_atual, row_fim)
            ws[celula_atual].value = 0
            coluna_atual = "H"
            celula_atual = "{}{}".format(coluna_atual, row_fim)
            ws[celula_atual].value = "=MAX(INDEX(GOOGLEFINANCE({}; \"high\"; $N$2; $O$2); ; 2))".format(cell_name_ativo)
            row_fim += 1
        
        print("Lista Vendidos")
        for item_vendido in list_vendidos:
            cell_name_desc = "{}{}".format("J", list_ativos_planilha.index(item_vendido) + 2)
            if ws[cell_name_desc].value is None or ws[cell_name_desc].value.find("Vendido pela Suno") == -1:
                ws[cell_name_desc] = "Vendido pela Suno em " + datetime.today().strftime('%d/%m/%Y')
                logging.info(str(datetime.now()) + ": FII " + item_vendido + " foi vendido")
                print(str(datetime.now()) + ": FII " + item_vendido + " foi vendido")
            ws[cell_name_desc].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        ws["{}{}".format("A", row_fim)] = "FIM"


        wb.save(file_name)
    except Exception as e:
        logging.info(str(datetime.now()) + ": Erro no update_worksheet: " + str(e))
        raise e
    

def main():
    logging.basicConfig(filename='/Users/marco/Documents/Logs/AtualizaAxelrodLog.txt', level=logging.INFO, filemode='w')
    logging.info(str(datetime.now()) + ": Achou que não ia ter programa hoje? Achou errado otário!")
    print(str(datetime.now()) + ": Achou que não ia ter programa hoje? Achou errado otário!")
    try:
        price_list_acao = [('EGIE3', 50.0), ('ENBR3', 25.0), ('VIVT3', 53.0), ('ITSA4', 12.0), ('ELET3', 41.0), ('LEVE3', 21.0), ('ABCB4', 18.5), ('BBAS3', 36.25), ('TAEE11', 35.0), ('GRND3', 8.33), ('WIZS3', 9.12), ('BBSE3', 33.75), ('PETR4', 30.0), ('ODPV3', 14.2), ('UNIP6', 35.0), ('CCRO3', 12.45), ('TUPY3', 20.0), ('SULA11', 48.0), ('ITSA4', 12.0), ('PNVL3', 27.0), ('MYPK3', 21.0), ('ALUP11', 31.0), ('MDIA3', 40.0), ('NEOE3', 19.0), ('EQTL3', 24.0), ('IGTA3', 35.0), ('GMAT3', 9.0), ('RLOG3', 23.0), ('CVCB3', 30.0), ('TFCO4', 13.35), ('CGRA4', 28.0), ('BRAP4', 58.0), ('BPAC11', 75.0), ('RENT3', 42.0)]
        price_list_fii = [('RBRP11', 100.0), ('MCCI11', 108.0), ('HGRU11', 135.0), ('BTLG11', 120.0), ('MALL11', 110.0), ('HSML11', 105.0), ('VILG11', 130.0), ('XPLG11', 125.0), ('ALZR11', 135.0), ('VGIP11', 115.0), ('VRTA11', 115.0), ('HGLG11', 170.0), ('IRDM11', 120.0), ('VISC11', 110.0), ('XPML11', 105.0), ('FIIB11', 525.0)]
        print(str(datetime.now()) + ": " + ("Tá rodando o debug animal" if rodar_debug else "Rodando certo finalmente ein"))
        if not rodar_debug:
            price_list_acao, price_list_fii = get_data()
        #price_list = logging.info(str(datetime.now()) + ": Cabô getData")
        print(str(datetime.now()) + ": Cabô getData")
        print(str(datetime.now()) + ": Lista de ativos:   " + str(price_list_acao))
        print(str(datetime.now()) + ": Lista de FII:   " + str(price_list_fii))
        logging.info(str(datetime.now()) + ": Lista de ativos:   " + str(price_list_acao))
        logging.info(str(datetime.now()) + ": Lista de ativos:   " + str(price_list_fii))
        update_worksheet(list(set(price_list_acao)), list(set(price_list_fii)))
        print(str(datetime.now()) + ": Cabô planilha")
        logging.info(str(datetime.now()) + ": Cabô planilha")
        logging.info(str(datetime.now()) + ": Cabô! Cabô o programa")
        print(str(datetime.now()) + ": Cabô! Cabô o programa")
    except:
        logging.info(str(datetime.now()) + ": Cabô com erro...")
        print(str(datetime.now()) + ": Cabô com erro...")


main()